"""
Export do Excelu

pandas umí uložit data i do Excel sešitu, se kterým se bude uživatelům lépe pracovat.
K ukádání do Excelu využívá pandas modul openpyxl, který ale není nainstalován automaticky.

    *   Nainstaluj si modul openpyxl standardním způsobem, který jsme si ukazovali v lekci.

    *    Ulož tabulku s cekovými počty vykázaných hodin, kterou jsi vytvořila v příkladu 27 jako Excel soubor.
        Pokud jsi tento příklad nezpracovala, ulož libovolnou jinou tabulku jako Excel sešit. Výsledný sešit si otevři a zkontroluj.
        K uložení použij funkci to_excel, se kterou pracuj stejně, jako jsme na lekci pracovali s funkci to_csv.

    *    Zkus data z Excelu naopak načíst pomocí funkci read_excel() a ověř, že se soubor načetl v pořádku.

"""
import pandas
from openpyxl import Workbook
from openpyxl.styles import PatternFill

#tabulku jsem vyexportoval do excelu v příkladu 27  --> priklad27.py

zamestnanci_vykazy = pandas.read_excel("zamestnanci_vykazy.xlsx")
print(zamestnanci_vykazy.head())


"""
ROZŠÍŘENÉ ZADÁNÍ

Modul openpyxl lze používat i samostatně. Přehled jeho funkcí najdeš v dokumentaci.

Zkus pomocí modulu zapsat rozvrh hodin jako tabulku v Excel sešitu. Níže máš program,
který obsahuje rozvrh hodin jako dvourozměrný seznam. Vnitřní seznamy obsahují předměty v rozvrhu, jeden vnitřní seznam vždy obsahuje předměty pro daný den.

Podívej se nejprve na ukázku. Jednoduchý program níže zapíše hodnotu Test do buňky B1 a výsledek uloží souboru rozvrh_hodin.xlsx.
Pokud neznáš terminologii Excelu, pak Workbook označuje sešit, tj. celý "soubor". ws se odkazuje na Work Sheet, což je list, tj. jedna "záložka".
Náš soubor bude mít jen jeden list, čímž je situace jednoduchá.

"""

"""
UKÁZKA:

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"
cislo_radku = 1
cislo_sloupce = 2

bunka = ws1.cell(cislo_radku, cislo_sloupce)
bunka.value = "Test"

wb.save(filename="rozvrh_hodin.xlsx")
"""

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(2, 1)
    bunka.value = #TOHLE FAKT NEVÍM :( --> napadá mě napsat ručně "Anglický jazyk" - ale určitě se s rozvrhem dá pracovat jinak...
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")